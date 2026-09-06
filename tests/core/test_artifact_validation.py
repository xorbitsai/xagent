"""Format checks are composable and never conflate execution with delivery."""

import hashlib
import json
import logging
import os
import re
import struct
import subprocess
import sys
from io import BytesIO
from unittest.mock import Mock
from zipfile import ZIP_DEFLATED, ZipFile

import pytest

from xagent.core.artifact_validation import (
    ArtifactCheck,
    ArtifactCheckRegistry,
    ArtifactContent,
    ValidationLimits,
    service,
)
from xagent.core.artifact_validation.defaults import default_registry
from xagent.core.artifact_validation.models import InvalidArtifact, UncheckedArtifact
from xagent.core.artifact_validation.service import validate_artifact


def check(data, filename, **limits):
    return default_registry().validate(
        ArtifactContent(filename, data, ValidationLimits(**limits))
    )


@pytest.mark.parametrize("extension", ["xlsx", "docx", "pptx", "pdf", "png"])
def test_corrupt_formats_are_invalid(extension):
    assert check(b"not a document", f"report.{extension}").status == "invalid"


@pytest.mark.parametrize(
    "extension", ["xlsx", "docx", "pptx", "pdf", "png", "bmp", "tif", "tiff"]
)
def test_real_readers_accept_minimal_documents(extension):
    stream = BytesIO()
    if extension == "xlsx":
        from openpyxl import Workbook

        Workbook().save(stream)
    elif extension == "docx":
        from docx import Document

        Document().save(stream)
    elif extension == "pptx":
        from pptx import Presentation

        Presentation().save(stream)
    elif extension == "pdf":
        from pypdf import PdfWriter

        writer = PdfWriter()
        writer.add_blank_page(width=100, height=100)
        writer.write(stream)
    else:
        from PIL import Image

        Image.new("RGB", (4, 4)).save(
            stream,
            format={"tif": "TIFF", "tiff": "TIFF"}.get(extension, extension.upper()),
        )
    result = check(stream.getvalue(), f"file.{extension}")
    assert result.status == "valid", result


@pytest.mark.parametrize(
    "extension,part",
    [
        ("xlsx", "xl/workbook.xml"),
        ("docx", "word/document.xml"),
        ("pptx", "ppt/presentation.xml"),
    ],
)
@pytest.mark.parametrize("missing", [False, True])
def test_office_reader_resolves_main_part(extension, part, missing):
    from xagent.core.artifact_validation.office import check_office_document

    original = BytesIO()
    if extension == "xlsx":
        from openpyxl import Workbook

        Workbook().save(original)
    elif extension == "docx":
        from docx import Document

        Document().save(original)
    else:
        from pptx import Presentation

        Presentation().save(original)
    renamed = part.replace(".xml", "-renamed.xml")
    old_rels = part.rsplit("/", 1)[0] + "/_rels/" + part.rsplit("/", 1)[1] + ".rels"
    new_rels = old_rels.replace(".xml.rels", "-renamed.xml.rels")
    rewritten = BytesIO()
    with ZipFile(original) as source, ZipFile(rewritten, "w", ZIP_DEFLATED) as target:
        for entry in source.infolist():
            if missing and entry.filename == part:
                continue
            data = source.read(entry)
            name = entry.filename
            if not missing:
                if name in ("[Content_Types].xml", "_rels/.rels"):
                    data = data.replace(part.encode(), renamed.encode())
                name = {part: renamed, old_rels: new_rels}.get(name, name)
            target.writestr(name, data)
    data = rewritten.getvalue()
    if not missing:
        # Prove that the installed format reader accepts the renamed package.
        check_office_document(
            ArtifactContent("file." + extension, data, ValidationLimits())
        )
    report = check(data, "file." + extension)
    assert [c.status for c in report.checks] == [
        "valid",
        "invalid" if missing else "valid",
    ]


@pytest.mark.parametrize("extension,format_name", [("gif", "GIF"), ("tiff", "TIFF")])
def test_image_checks_all_frames(extension, format_name, monkeypatch):
    from PIL import Image

    stream = BytesIO()
    Image.new("RGB", (4, 4), "red").save(
        stream,
        format=format_name,
        save_all=True,
        append_images=[Image.new("RGB", (4, 4), "blue")],
    )
    data = stream.getvalue()
    assert check(data, "frames." + extension).status == "valid"
    assert check(data, "frames." + extension, max_pixels=31).status == "unchecked"
    with Image.open(BytesIO(data)) as image:
        reader_class = type(image)
    original_load = reader_class.load

    def fail_second_frame(self, *args, **kwargs):
        if self.tell() == 1:
            raise OSError("second frame cannot decode")
        return original_load(self, *args, **kwargs)

    monkeypatch.setattr(reader_class, "load", fail_second_frame)
    assert check(data, "frames." + extension).status == "invalid"


def test_read_only_xlsx_checks_cells_beyond_stale_dimensions():
    from openpyxl import Workbook

    original = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["header"])
    sheet.append([1])
    sheet.append([123])
    workbook.save(original)

    rewritten = BytesIO()
    with ZipFile(original) as source, ZipFile(rewritten, "w", ZIP_DEFLATED) as target:
        for entry in source.infolist():
            data = source.read(entry)
            if entry.filename == "xl/worksheets/sheet1.xml":
                assert b'<dimension ref="A1:A3"/>' in data
                assert b"<v>123</v>" in data
                data = data.replace(
                    b'<dimension ref="A1:A3"/>', b'<dimension ref="A1:A1"/>'
                )
                data = data.replace(b"<v>123</v>", b"<v>not-a-number</v>")
            target.writestr(entry, data)

    # The XML/ZIP remains well formed. Without resetting the read-only sheet's
    # bounds, iteration stops at row 2 before decoding the corrupt third row.
    report = check(rewritten.getvalue(), "stale-dimensions.xlsx")
    assert [c.status for c in report.checks] == ["valid", "invalid"]


@pytest.mark.parametrize(
    "data",
    [
        b"",
        b"a,b\n1,2\n",
        b"one column",
        b"a;b\n1;2",
        b"a,b\n1,2,3",
        "名字,值\n甲,1".encode("utf-16"),
    ],
)
def test_csv_does_not_impose_business_schema(data):
    assert check(data, "data.CSV").status == "valid"


def test_csv_malformed_encoding_and_budget():
    assert check(b'a,b\n"unterminated,2', "data.csv").status == "unchecked"
    assert check(b"\xff\x98", "data.csv").status == "unchecked"
    assert check(b"a,\x00", "data.csv").status == "invalid"
    assert check(b"a\nb", "data.csv", max_units=1).status == "unchecked"
    assert check(b"one", "data.csv", max_bytes=2).status == "unchecked"
    assert check(b"whatever", "data.unknown").status == "unchecked"


def test_csv_lenient_reader_recovery_is_not_reported_as_corruption():
    import csv

    data = b'a,b\n1,"he said "hi" there"\n'
    assert len(list(csv.reader(data.decode().splitlines()))) == 2
    assert check(data, "data.csv").status == "unchecked"


@pytest.mark.parametrize(
    "existing",
    [(-1, -1), (512 * 1024**2, 512 * 1024**2), (256 * 1024**2, 512 * 1024**2)],
)
def test_worker_keeps_stricter_address_space_limits(existing, monkeypatch):
    from io import StringIO
    from types import SimpleNamespace

    from xagent.core.artifact_validation import worker

    resource = pytest.importorskip("resource")
    monkeypatch.setattr(resource, "RLIM_INFINITY", -1)
    applied = []

    def set_limit(kind, limits):
        assert kind == resource.RLIMIT_AS
        if existing[1] != -1 and limits[1] > existing[1]:
            raise ValueError("cannot raise the hard limit")
        applied.append(limits)

    monkeypatch.setattr(resource, "getrlimit", lambda _: existing)
    monkeypatch.setattr(resource, "setrlimit", set_limit)
    monkeypatch.setattr(sys, "platform", "linux")
    monkeypatch.setattr(sys, "argv", ["worker", "data.csv", "1024"])
    monkeypatch.setattr(sys, "stdin", SimpleNamespace(buffer=BytesIO(b"a,b\n1,2\n")))
    output = StringIO()
    flush = Mock(wraps=output.flush)
    monkeypatch.setattr(output, "flush", flush)
    monkeypatch.setattr(sys, "stdout", output)
    worker.main()
    flush.assert_called_once_with()
    assert applied == [
        tuple(1024**3 if limit == -1 else min(limit, 1024**3) for limit in existing)
    ]
    assert json.loads(output.getvalue())["status"] == "valid"


def test_archive_budgets_and_unsafe_xml():
    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as archive:
        for name in ["[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"]:
            archive.writestr(name, "<a/>" * 20)
    assert (
        check(stream.getvalue(), "data.xlsx", max_expanded_bytes=10).status
        == "unchecked"
    )
    stream = BytesIO()
    with ZipFile(stream, "w") as archive:
        for name in ["[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"]:
            archive.writestr(name, '<!DOCTYPE a [<!ENTITY x "value">]><a>&x;</a>')
    assert check(stream.getvalue(), "data.xlsx").status == "invalid"


@pytest.mark.parametrize(
    "suffix,root",
    [
        ("xlsx", "xl/workbook.xml"),
        ("docx", "word/document.xml"),
        ("pptx", "ppt/presentation.xml"),
    ],
)
@pytest.mark.parametrize(
    "bad_member",
    ["/absolute.xml", "../outside.xml", "xl/../../outside.xml", "[Content_Types].xml"],
)
def test_office_package_rejects_ambiguous_members(suffix, root, bad_member):
    stream = BytesIO()
    with ZipFile(stream, "w") as archive:
        for name in ["[Content_Types].xml", "_rels/.rels", root, bad_member]:
            archive.writestr(name, "<a/>")
    report = check(stream.getvalue(), f"file.{suffix}")
    assert report.status == "invalid"
    assert len(report.checks) == 1
    assert "ambiguous member paths" in report.checks[0].message


@pytest.mark.parametrize("missing", ["[Content_Types].xml", "_rels/.rels"])
def test_office_package_rejects_missing_required_parts(missing):
    stream = BytesIO()
    with ZipFile(stream, "w") as archive:
        for name in ["[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"]:
            if name != missing:
                archive.writestr(name, "<a/>")
    report = check(stream.getvalue(), "file.xlsx")
    assert report.status == "invalid"
    assert "missing required document parts" in report.checks[0].message


def test_office_package_entry_count_and_encrypted_member_guards():
    stream = BytesIO()
    with ZipFile(stream, "w") as archive:
        for name in ["[Content_Types].xml", "_rels/.rels", "xl/workbook.xml"]:
            archive.writestr(name, "<a/>")
    report = check(stream.getvalue(), "file.xlsx", max_entries=2)
    assert report.status == "unchecked"
    assert "expansion budget" in report.checks[0].message

    data = bytearray(stream.getvalue())
    # Mark the first local and central-directory entry encrypted. The guard
    # must stop before attempting to decode its untrusted payload.
    struct.pack_into("<H", data, data.index(b"PK\x03\x04") + 6, 1)
    struct.pack_into("<H", data, data.index(b"PK\x01\x02") + 8, 1)
    report = check(bytes(data), "file.xlsx")
    assert report.status == "unchecked"
    assert "Encrypted Office" in report.checks[0].message


@pytest.mark.parametrize(
    "payload,expected", [(b"", "valid"), (b"q Q\n", "valid"), (None, "unchecked")]
)
@pytest.mark.parametrize("array", [False, True])
def test_pdf_does_not_certify_silently_discarded_flate_content(
    payload, expected, array
):
    import zlib

    from pypdf import PdfWriter
    from pypdf.generic import ArrayObject, EncodedStreamObject, NameObject

    writer = PdfWriter()
    page = writer.add_blank_page(width=100, height=100)
    stream = EncodedStreamObject()
    stream[NameObject("/Filter")] = NameObject("/FlateDecode")
    stream._data = (
        zlib.compress(payload) if payload is not None else b"not compressed data"
    )
    # A checksum defect can still recover every original content byte. Do not
    # reject successful nonempty recovery just for strict zlib conformance.
    if payload:
        stream._data = stream._data[:-1] + bytes([stream._data[-1] ^ 1])
        assert stream.get_data() == payload
    ref = writer._add_object(stream)
    page[NameObject("/Contents")] = ArrayObject([ref]) if array else ref
    output = BytesIO()
    writer.write(output)
    report = check(output.getvalue(), "content.pdf")
    assert report.status == expected
    if expected == "unchecked":
        assert "stream recovery" in report.checks[0].message


@pytest.mark.parametrize(
    "pages,max_units,expected",
    [(500, 200_000, "valid"), (501, 200_000, "unchecked"), (3, 2, "unchecked")],
)
def test_pdf_page_ceiling_is_not_partial_success(pages, max_units, expected):
    from pypdf import PdfWriter

    writer = PdfWriter()
    for _ in range(pages):
        writer.add_blank_page(width=10, height=10)
    output = BytesIO()
    writer.write(output)
    report = check(output.getvalue(), "pages.pdf", max_units=max_units)
    assert report.status == expected
    if expected == "unchecked":
        assert f"{min(max_units, 500)} pages" in report.checks[-1].message


@pytest.mark.parametrize("error_name", ["LimitReachedError", "DependencyError"])
def test_pdf_reader_limits_and_dependencies_are_expected_unchecked(
    monkeypatch, caplog, error_name
):
    import pypdf
    from pypdf import errors

    error_type = getattr(errors, error_name)
    monkeypatch.setattr(
        pypdf, "PdfReader", Mock(side_effect=error_type("private parser detail"))
    )
    report = check(b"%PDF-1.7\n", "limited.pdf")
    assert report.status == "unchecked"
    assert "dependency or parser limit" in report.checks[-1].message
    assert "private parser detail" not in str(report.as_dict())
    assert not caplog.records


def test_pdf_missing_encoded_stream_api_has_explicit_diagnostic(monkeypatch):
    from types import SimpleNamespace

    import pypdf
    from pypdf.generic import EncodedStreamObject, NameObject

    stream = EncodedStreamObject()
    stream[NameObject("/Filter")] = NameObject("/FlateDecode")
    stream._data = b""
    del stream._data
    monkeypatch.setattr(stream, "get_data", lambda: b"")
    monkeypatch.setattr(
        pypdf,
        "PdfReader",
        lambda *args, **kwargs: SimpleNamespace(
            is_encrypted=False, pages=[{"/Contents": stream}]
        ),
    )
    report = check(b"%PDF-1.7\n", "stream.pdf")
    assert report.status == "unchecked"
    assert "does not expose encoded stream bytes" in report.checks[-1].message


def test_readable_pdf_with_recoverable_xref_is_not_invalid():
    from pypdf import PdfReader, PdfWriter
    from pypdf.errors import PdfReadError

    stream = BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=100, height=100)
    writer.write(stream)
    original = stream.getvalue()
    match = re.search(rb"startxref\s+(\d+)", original)
    assert match
    data = (
        original[: match.start(1)]
        + str(int(match[1]) + 1).encode()
        + original[match.end(1) :]
    )
    with pytest.raises(PdfReadError):
        PdfReader(BytesIO(data), strict=True)
    assert len(PdfReader(BytesIO(data), strict=False).pages) == 1
    assert check(data, "recoverable.pdf").status == "valid"
    assert check(b"%PDF-1.7\nnot a document", "broken.pdf").status == "invalid"


def test_cache_keeps_filename_and_extension_with_identical_bytes(tmp_path, monkeypatch):
    service._cache.clear()
    path = tmp_path / "artifact.bin"
    path.write_bytes(b"a,b\n1,2")
    run = Mock(wraps=service._run_checks)
    monkeypatch.setattr(service, "_run_checks", run)
    for filename, status in [
        ("one.csv", "valid"),
        ("two.csv", "valid"),
        ("one.pdf", "invalid"),
    ]:
        assert validate_artifact(path, filename=filename).status == status
        assert validate_artifact(path, filename=filename).status == status
    assert run.call_count == 3


def test_image_pixel_budget_and_pdf_encryption():
    from PIL import Image
    from pypdf import PdfWriter

    stream = BytesIO()
    Image.new("RGB", (4, 4)).save(stream, format="PNG")
    assert check(stream.getvalue(), "image.png", max_pixels=4).status == "unchecked"
    stream = BytesIO()
    writer = PdfWriter()
    writer.add_blank_page(width=10, height=10)
    writer.encrypt("password")
    writer.write(stream)
    assert check(stream.getvalue(), "private.pdf").status == "unchecked"


@pytest.mark.parametrize(
    ("error", "status"),
    [
        (InvalidArtifact("bad structure"), "invalid"),
        (UncheckedArtifact("budget"), "unchecked"),
        (ImportError("missing"), "unchecked"),
        (RuntimeError("/secret/path"), "unchecked"),
    ],
)
def test_additional_checks_and_fail_closed_preflight(error, status):
    registry = ArtifactCheckRegistry()
    first = Mock()
    second = Mock(side_effect=error)
    third = Mock()
    for name, callback in [("first", first), ("second", second), ("third", third)]:
        registry.register(ArtifactCheck(name, frozenset({".foo"}), callback))
    result = registry.validate(ArtifactContent("new.foo", b"bytes", ValidationLimits()))
    assert result.status == status
    assert [c.name for c in result.checks] == ["first", "second"]
    assert "/secret/path" not in str(result.as_dict())
    third.assert_not_called()
    with pytest.raises(ValueError):
        registry.register(ArtifactCheck("first", frozenset({".foo"}), first))


def test_service_real_worker_and_same_name_rewrite(tmp_path):
    path = tmp_path / "test.csv"
    path.write_bytes(b"a,b\n1,2")
    first = validate_artifact(path)
    assert first.status == "valid"
    assert first.sha256 == hashlib.sha256(path.read_bytes()).hexdigest()
    path.write_bytes(b'a,b\n"broken')
    second = validate_artifact(path)
    assert second.status == "unchecked"
    assert second.sha256 != first.sha256
    assert path.read_bytes() == b'a,b\n"broken'


@pytest.mark.parametrize(
    "error", [MemoryError("read failed"), RuntimeError("unexpected read failure")]
)
def test_snapshot_exception_does_not_drop_generated_attachment(
    tmp_path, monkeypatch, caplog, error
):
    from contextlib import contextmanager
    from types import SimpleNamespace

    from xagent.core.tools.artifacts import build_generated_file_metadata

    path = tmp_path / "output.csv"
    path.write_bytes(b"a,b\n1,2\n")
    original = service.os.fdopen

    @contextmanager
    def failing_open(*args, **kwargs):
        with original(*args, **kwargs) as stream:
            wrapped = Mock(wraps=stream)
            wrapped.read.side_effect = error
            yield wrapped

    monkeypatch.setattr(service.os, "fdopen", failing_open)
    workspace = SimpleNamespace(
        workspace_dir=tmp_path, get_file_id_from_path=lambda _: "registered-file"
    )
    metadata = build_generated_file_metadata(workspace=workspace, file_paths=[path])
    assert len(metadata["file_refs"]) == 1
    assert metadata["file_refs"][0]["file_id"] == "registered-file"
    assert metadata["file_refs"][0]["validation"]["status"] == "unchecked"
    assert path.read_bytes() == b"a,b\n1,2\n"
    assert "snapshot validation failed" in caplog.text
    assert service._slots.acquire(blocking=False)
    assert service._slots.acquire(blocking=False)
    service._slots.release()
    service._slots.release()


def test_cache_is_byte_keyed_and_detects_mid_check_rewrite(tmp_path, monkeypatch):
    service._cache.clear()
    path = tmp_path / "test.csv"
    path.write_bytes(b"a,b")
    run = Mock(wraps=service._run_checks)
    monkeypatch.setattr(service, "_run_checks", run)
    assert validate_artifact(path).status == "valid"
    assert validate_artifact(path).status == "valid"
    assert run.call_count == 1
    service._cache.clear()

    def rewrite(filename, data, *_args):
        path.write_bytes(b"changed")
        return check(data, filename)

    monkeypatch.setattr(service, "_run_checks", rewrite)
    assert validate_artifact(path).status == "unchecked"


def test_worker_failure_timeout_missing_file_and_byte_budget(tmp_path, monkeypatch):
    service._cache.clear()
    path = tmp_path / "test.csv"
    path.write_bytes(b"abc")
    monkeypatch.setattr(
        service.subprocess,
        "run",
        Mock(side_effect=subprocess.TimeoutExpired("validator", 1)),
    )
    assert validate_artifact(path).status == "unchecked"
    monkeypatch.setenv("XAGENT_ARTIFACT_VALIDATION_MAX_BYTES", "2")
    assert validate_artifact(path).status == "unchecked"
    assert validate_artifact(tmp_path / "missing.csv").status == "unchecked"


@pytest.mark.parametrize(
    "error",
    [OSError("server-only detail"), subprocess.CalledProcessError(1, ["validator"])],
)
def test_worker_failures_are_logged_without_exposing_details(
    monkeypatch, caplog, error
):
    monkeypatch.setattr(service.subprocess, "run", Mock(side_effect=error))
    with caplog.at_level(logging.ERROR, logger=service.__name__):
        report = service._run_checks("data.csv", b"a,b", 1024, 1)

    assert report.status == "unchecked"
    assert report.checks[0].message == "Validator process could not complete."
    assert "server-only detail" not in str(report.as_dict())
    record = caplog.records[-1]
    assert record.name == service.__name__
    assert record.exc_info[1] is error


def test_malformed_worker_response_retains_server_trace(monkeypatch, caplog):
    monkeypatch.setattr(
        service.subprocess,
        "run",
        Mock(
            return_value=subprocess.CompletedProcess(
                ["validator"], 0, stdout=b"private payload"
            )
        ),
    )
    with caplog.at_level(logging.ERROR, logger=service.__name__):
        report = service._run_checks("data.csv", b"a,b", 1024, 1)

    assert report.status == "unchecked"
    assert isinstance(caplog.records[-1].exc_info[1], json.JSONDecodeError)
    assert "private payload" not in str(report.as_dict())


def test_real_worker_preserves_binary_stdin_bytes():
    # CRLF, Ctrl-Z, NUL and every byte value catch text-mode translation or
    # truncation. Exercise the actual parent/worker pipe, not a mocked stream.
    payload = b"PK\x03\x04\r\n\x1a\x00\xff" + bytes(range(256)) * 4
    report = service._run_checks("payload.bin", payload, len(payload), 8)
    assert report.status == "unchecked"  # No format reader is selected.
    assert report.sha256 == hashlib.sha256(payload).hexdigest()


def test_sandbox_cannot_assert_host_validation(tmp_path, monkeypatch):
    monkeypatch.setattr(service, "in_sandbox_tool_runner", lambda: True)
    assert validate_artifact(tmp_path / "data.csv").status == "unchecked"


def test_real_office_dependency_absence_is_unchecked(monkeypatch):
    from openpyxl import Workbook

    stream = BytesIO()
    Workbook().save(stream)
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    report = check(stream.getvalue(), "file.xlsx")
    assert report.status == "unchecked"
    assert [c.status for c in report.checks] == ["valid", "unchecked"]


def test_invalid_configuration_preserves_an_unchecked_report(
    tmp_path, monkeypatch, caplog
):
    monkeypatch.setenv("XAGENT_ARTIFACT_VALIDATION_MAX_BYTES", "invalid")
    assert validate_artifact(tmp_path / "file.csv").status == "unchecked"
    assert "Invalid artifact validation configuration" in caplog.text


def test_public_validation_cannot_exhaust_private_capacity(tmp_path, monkeypatch):
    from concurrent.futures import ThreadPoolExecutor
    from threading import BoundedSemaphore, Event

    service._cache.clear()
    monkeypatch.setattr(service, "_slots", BoundedSemaphore(2))
    monkeypatch.setattr(service, "_public_slots", BoundedSemaphore(1))
    entered, release = Event(), Event()
    path = tmp_path / "data.csv"
    path.write_bytes(b"a,b")

    def slow_public(filename, data, *_args):
        if filename == "public.csv":
            entered.set()
            assert release.wait(5)
        return check(data, filename)

    monkeypatch.setattr(service, "_run_checks", slow_public)
    with ThreadPoolExecutor(max_workers=1) as pool:
        public = pool.submit(
            validate_artifact, path, filename="public.csv", public=True
        )
        try:
            assert entered.wait(5)
            assert validate_artifact(path, public=True).status == "unchecked"
            assert validate_artifact(path).status == "valid"
        finally:
            release.set()
        assert public.result(timeout=5).status == "valid"
    assert validate_artifact(path, public=True).status == "valid"


def test_public_validation_releases_capacity_on_unexpected_failure(
    tmp_path, monkeypatch
):
    from threading import BoundedSemaphore

    public = BoundedSemaphore(1)
    monkeypatch.setattr(service, "_public_slots", public)
    monkeypatch.setattr(
        service, "_validate_snapshot", Mock(side_effect=RuntimeError("failed"))
    )
    assert validate_artifact(tmp_path / "data.csv", public=True).status == "unchecked"
    assert public.acquire(blocking=False)
    public.release()


@pytest.mark.skipif(not hasattr(os, "mkfifo"), reason="FIFO not available")
def test_non_regular_files_do_not_block_snapshot_reads(tmp_path):
    path = tmp_path / "pipe.csv"
    os.mkfifo(path)
    assert validate_artifact(path).status == "unchecked"
